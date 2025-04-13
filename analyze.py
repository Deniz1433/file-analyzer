#!/usr/bin/env python3
import os
import sys
import time
import re
import math
import magic
import zipfile
import xml.etree.ElementTree as ET
import rarfile
import py7zr
import msoffcrypto
import oletools.olevba as olevba
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import pefile
import subprocess
import pypackerdetect
import json
import concurrent.futures

# Define common TLDs and compile the domain regex
COMMON_TLDS = (
    "com|org|net|edu|gov|mil|int|arpa|biz|info|name|pro|aero|coop|museum|"
    "ac|ad|ae|af|ag|ai|al|am|ao|aq|ar|as|at|au|aw|ax|az|ba|bb|bd|be|bf|bg|bh|bi|bj|bm|bn|bo|br|bs|bt|bv|bw|by|bz|"
    "ca|cat|cc|cd|cf|cg|ch|ci|ck|cl|cm|cn|co|cr|cu|cv|cw|cx|cy|cz|"
    "de|dj|dk|dm|do|dz|ec|ee|eg|er|es|et|eu|"
    "fi|fj|fk|fm|fo|fr|"
    "ga|gb|gd|ge|gf|gg|gh|gi|gl|gm|gn|gp|gq|gr|gs|gt|gu|gw|gy|"
    "hk|hm|hn|hr|ht|hu|"
    "id|ie|il|im|in|io|iq|ir|is|it|"
    "je|jm|jo|jobs|jp|"
    "ke|kg|kh|ki|km|kn|kp|kr|kw|ky|kz|"
    "la|lb|lc|li|lk|lr|ls|lt|lu|lv|ly|"
    "ma|mc|md|me|mg|mh|mil|mk|ml|mm|mn|mo|mobi|mp|mq|mr|ms|mt|mu|museum|mv|mw|mx|my|mz|"
    "na|name|nc|ne|net|nf|ng|ni|nl|no|np|nr|nu|nz|"
    "om|org|pa|pe|pf|pg|ph|pk|pl|pm|pn|post|pr|pro|ps|pt|pw|py|"
    "qa|re|ro|rs|ru|rw|"
    "sa|sb|sc|sd|se|sg|sh|si|sk|sl|sm|sn|so|sr|st|su|sv|sx|sy|sz|"
    "tc|td|tel|tf|tg|th|tj|tk|tl|tm|tn|to|tp|tr|travel|tt|tv|tw|tz|"
    "ua|ug|uk|us|uy|uz|"
    "va|vc|ve|vg|vi|vn|vu|"
    "wf|ws|xxx|ye|yt|za|zm|zw"
)
DOMAIN_REGEX = re.compile(
    r'\b(?:[a-zA-Z0-9-]+\.)+(?:' + COMMON_TLDS + r')\b', re.IGNORECASE
)
ALLOWED_LANGS = {
    'ar-SA', 'bg-BG', 'zh-CN', 'zh-TW', 'hr-HR', 'cs-CZ', 'da-DK', 'nl-NL',
    'en-US', 'et-EE', 'fi-FI', 'fr-FR', 'de-DE', 'el-GR', 'he-IL', 'hi-IN',
    'hu-HU', 'id-ID', 'it-IT', 'ja-JP', 'kk-KZ', 'ko-KR', 'lv-LV', 'lt-LT',
    'ms-MY', 'nb-NO', 'pl-PL', 'pt-BR', 'pt-PT', 'ro-RO', 'ru-RU', 'sr-latn-RS',
    'sk-SK', 'sl-SI', 'es-ES', 'sv-SE', 'th-TH', 'tr-TR', 'uk-UA', 'vi-VN'
}


def compute_entropy(data):
    if not data:
        return 0.0
    frequency = {}
    for b in data:
        frequency[b] = frequency.get(b, 0) + 1
    entropy = 0.0
    for count in frequency.values():
        p = count / len(data)
        entropy -= p * math.log2(p)
    return entropy


def analyze_file(file_path):
    with open(file_path, 'rb') as f:
        data = f.read(2048)
    return magic.from_buffer(data), magic.from_buffer(data, mime=True)


def is_office_file(desc, mime):
    indicators = [
        'microsoft word', 'word document', 'ms word',
        'microsoft excel', 'excel',
        'microsoft power point', 'powerpoint',
        'office open xml', 'ms-office'
    ]
    return any(ind in desc.lower() or ind in mime.lower() for ind in indicators)


def check_archive_password(file_path, desc):
    d = desc.lower()
    if "zip" in d:
        try:
            with zipfile.ZipFile(file_path) as z:
                return any(info.flag_bits & 0x1 for info in z.infolist())
        except Exception:
            return False
    elif "rar" in d:
        try:
            rf = rarfile.RarFile(file_path)
            rf.testrar()
            return False
        except rarfile.RarPasswordNeeded:
            return True
        except Exception:
            return False
    elif "7-zip" in d or "7z" in d:
        try:
            with py7zr.SevenZipFile(file_path, mode='r') as archive:
                archive.getnames()
            return False
        except py7zr.exceptions.PasswordRequired:
            return True
        except Exception:
            return False
    return False


def analyze_pdf(file_path):
    try:
        reader = PdfReader(file_path)
    except Exception as e:
        sys.stderr.write("Error reading PDF: " + str(e) + "\n")
        return None, set(), set(), set()
    encrypted = reader.is_encrypted
    urls, ips, domains = set(), set(), set()
    if not encrypted:
        text = " ".join(filter(None, (page.extract_text() for page in reader.pages)))
        urls.update(re.findall(r'\bhttps?://[^\s,]+', text))
        ips.update(re.findall(r'\b(?:\d{1,3}\.){3}\d{1,3}\b', text))
        domains.update(DOMAIN_REGEX.findall(text))
    return encrypted, urls, ips, domains


def extract_office_language(file_path, ext):
    def get_lang(xml_data, tag_match):
        try:
            root = ET.fromstring(xml_data)
            for elem in root.iter():
                if tag_match(elem):
                    lang_val = (elem.attrib.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val')
                                or elem.attrib.get('val') or (elem.text and elem.text.strip()))
                    if lang_val and lang_val in ALLOWED_LANGS:
                        return lang_val.strip()
        except ET.ParseError:
            pass
        return None

    candidates = []
    if ext in ['.docx', '.docm']:
        candidates = [
            ('word/settings.xml', lambda e: 'lang' in e.tag.lower()),
            ('docProps/core.xml', lambda e: e.tag.lower().endswith('language'))
        ]
    elif ext in ['.xlsx', '.xlsm']:
        candidates = [
            ('xl/theme/theme1.xml', lambda e: 'themefontlang' in e.tag.lower()),
            ('docProps/core.xml', lambda e: e.tag.lower().endswith('language'))
        ]
    try:
        with zipfile.ZipFile(file_path) as z:
            for fname, tag_fn in candidates:
                if fname in z.namelist():
                    lang = get_lang(z.read(fname), tag_fn)
                    if lang:
                        return lang
    except Exception as e:
        sys.stderr.write("Error extracting language: " + str(e) + "\n")
    return "Unknown"


def process_office_file(file_path, desc, mime):
    ext = os.path.splitext(file_path)[1].lower()
    try:
        with open(file_path, 'rb') as f:
            office = msoffcrypto.OfficeFile(f)
            encrypted = office.is_encrypted()
    except Exception:
        encrypted = False
    try:
        vba = olevba.VBA_Parser(file_path)
        has_macros = vba.detect_vba_macros()
    except Exception:
        has_macros = False
    language = extract_office_language(file_path, ext)
    page_count = "Unknown"
    if ext in ['.docx', '.docm']:
        try:
            with zipfile.ZipFile(file_path) as z:
                xml_data = z.read("word/document.xml").decode("utf-8")
            page_count = xml_data.count('w:type="page"') + 1
        except Exception:
            pass
    elif ext in ['.xlsx', '.xlsm']:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            page_count = len(wb.sheetnames)
        except Exception:
            pass
    elif ext in ['.pptx', '.pptm']:
        try:
            with zipfile.ZipFile(file_path) as z:
                slides = [name for name in z.namelist() 
                          if name.startswith("ppt/slides/slide") and name.endswith(".xml")]
                page_count = len(slides)
        except Exception:
            pass

    return {
        "encrypted": encrypted,
        "has_macros": has_macros,
        "language": language,
        "page_count": page_count
    }


def extract_ascii_strings(file_path, min_length=4):
    try:
        with open(file_path, 'rb') as f:
            data = f.read()
        return [s.decode('utf-8', errors='ignore')
                for s in re.findall(rb'[\x20-\x7E]{' + str(min_length).encode() + rb',}', data)]
    except Exception as e:
        sys.stderr.write("Error extracting strings from {}: {}\n".format(file_path, e))
        return []


def check_pe_header_info(file_path):
    details = {}
    try:
        pe = pefile.PE(file_path)
        machine = pe.FILE_HEADER.Machine
        if machine == 0x014C:
            arch = "x86"
        elif machine == 0x8664:
            arch = "x86-64"
        else:
            arch = hex(machine)
        details["architecture"] = arch
        details["file_size_bytes"] = os.path.getsize(file_path)
        with open(file_path, "rb") as f:
            file_data = f.read()
        details["general_entropy"] = round(compute_entropy(file_data), 4)
        sections = []
        for section in pe.sections:
            section_data = section.get_data()
            sec_entropy = round(compute_entropy(section_data), 4) if section_data else "N/A"
            sections.append({
                "name": section.Name.decode('utf-8', errors='ignore').strip('\x00'),
                "virtual_size": section.Misc_VirtualSize,
                "raw_size": section.SizeOfRawData,
                "entropy": sec_entropy
            })
        details["number_of_sections"] = len(pe.sections)
        details["sections"] = sections
        timestamp = pe.FILE_HEADER.TimeDateStamp
        details["compilation_date"] = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime(timestamp))
        if hasattr(pe, "DIRECTORY_ENTRY_IMPORT"):
            details["imported_dlls"] = [entry.dll.decode('utf-8', errors='ignore')
                                        for entry in pe.DIRECTORY_ENTRY_IMPORT]
        return details
    except Exception as e:
        sys.stderr.write("Error processing PE header for {}: {}\n".format(file_path, e))
        return {"error": str(e)}


def check_pe_details(file_path):
    details = {"urls": [], "domains": [], "ips": []}
    candidate_strings = extract_ascii_strings(file_path, min_length=4)
    url_pattern = re.compile(r'https?://[^\s\'"<>]+')
    ip_pattern = re.compile(r'\b(?:\d{1,3}\.){3}\d{1,3}\b')
    for s in candidate_strings:
        details["urls"].extend(url_pattern.findall(s))
        details["ips"].extend(ip_pattern.findall(s))
        details["domains"].extend(DOMAIN_REGEX.findall(s))
    details["urls"] = list(set(details["urls"]))
    details["ips"] = list(set(details["ips"]))
    details["domains"] = list(set(details["domains"]))
    return details


def filter_ips(ip_list):
    filtered = []
    for ip in ip_list:
        if sum(1 for octet in ip.split('.') if octet == "0") < 3:
            filtered.append(ip)
    return filtered


def check_packing(file_path):
    """
    Call pypackerdetect with --peid-large-db.
    - Ignore PEID signatures for Microsoft Visual C++, Visual C#, and Basic .NET when counting detections.
    - Count all other detections (including UPX PEID signatures) as packed.
    - Collect explicit packer matches and non-C++/C#/VB.NET PEID signatures as packer names.
    """
    try:
        cmd = ["pypackerdetect", "--peid-large-db", file_path]
        si = None
        if os.name == "nt":
            si = subprocess.STARTUPINFO()
            si.dwFlags |= subprocess.STARTF_USESHOWWINDOW

        result = subprocess.run(cmd, capture_output=True, text=True, startupinfo=si)
        output = result.stdout

        # 1) Gather all heuristic detection lines
        detection_lines = re.findall(r"\[DETECTION\]\s*(.*)", output)

        real_detections = []
        peid_sig_names = []

        for d in detection_lines:
            if d.startswith("Found PEID signature:"):
                # Extract the signature name
                m = re.match(r"Found PEID signature:\s*(.*?)\s*(?:->|$)", d)
                name = m.group(1).strip() if m else None
                if name and not any(skip in name for skip in ["Visual C++", "Visual C#", "Basic .NET"]):
                    real_detections.append(d)
                    peid_sig_names.append(name)
            else:
                # Any other detection counts
                real_detections.append(d)

        # 2) Explicit "matches known packer" entries
        packer_matches = re.findall(r"matches known packer:\s*\[(.*?)\]", output)

        # 3) Determine packed flag
        packed = bool(real_detections) or bool(packer_matches)

        # 4) Build packer name list
        packer_names = []
        if packer_matches:
            packer_names.extend(packer_matches)
        packer_names.extend(peid_sig_names)

        packer = None
        if packer_names:
            packer = ", ".join(sorted(set(packer_names)))

        return {
            "packed": packed,
            "packer": packer,
            "output": output
        }

    except Exception as e:
        return {"error": str(e)}


def analyze_file_path(file_path):
    results = {}
    results["file_path"] = file_path
    desc, mime = analyze_file(file_path)
    results["description"] = desc
    results["mime"] = mime

    if is_office_file(desc, mime):
        results["type"] = "office"
        results["office_details"] = process_office_file(file_path, desc, mime)

    elif "pdf" in mime.lower() or "pdf" in desc.lower():
        results["type"] = "pdf"
        pdf_enc, urls, ips, domains = analyze_pdf(file_path)
        results["pdf_details"] = {
            "encrypted": pdf_enc,
            "urls": list(urls),
            "ips": list(ips),
            "domains": list(domains)
        }

    # PE detection by magic, not extension
    elif (
        "executable" in desc.lower()
        or "pe32" in desc.lower()
        or "application/x-dosexec" in mime.lower()
    ):
        results["type"] = "pe"
        pe_header = check_pe_header_info(file_path)
        pe_det = check_pe_details(file_path)
        pe_det["ips"] = filter_ips(pe_det.get("ips", []))
        pack_result = check_packing(file_path)
        results["pe_header"] = pe_header
        results["pe_details"] = pe_det
        results["packing_detection"] = pack_result

    elif any(x in desc.lower() for x in ['zip', 'rar', '7-zip', '7z']):
        results["type"] = "archive"
        results["archive_protected"] = check_archive_password(file_path, desc)

    else:
        results["type"] = "other"

    return results



if __name__ == '__main__':
    try:
        input_text = sys.stdin.read()
        file_list = json.loads(input_text)
    except Exception as e:
        sys.stderr.write(f"Failed to parse input: {e}\n")
        sys.exit(1)

    n = len(file_list)

    for idx, entry in enumerate(file_list, start=1):
        file_path = entry["path"]
        orig_name = entry["name"]

        sys.stdout.write(f"PROGRESS {idx}/{n}\n")
        sys.stdout.flush()

        try:
            result = analyze_file_path(file_path)
        except Exception as e:
            result = {"error": str(e)}

        sys.stdout.write(json.dumps({orig_name: result}) + "\n")
        sys.stdout.flush()

    sys.stdout.write(f"DONE {n}/{n}\n")
    sys.stdout.flush()
